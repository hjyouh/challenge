import json
import re
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
INPUT = ROOT / "assets" / "data" / "challenge-attendance.xlsx"
OUTPUT = ROOT / "js" / "imported-data.js"


def sheet_year_month(name, fallback_order):
    digits = re.sub(r"\D", "", name)
    if digits.startswith("26") and len(digits) >= 3:
        return 2026, int(digits[2:])
    if digits.startswith("25") and len(digits) >= 3:
        return 2025, int(digits[2:])
    if digits:
        return 2025, int(digits[-2:] if len(digits) > 1 else digits)
    return 2025, fallback_order


def mission_key(year, label):
    if not isinstance(label, str):
        return None
    match = re.search(r"(\d{1,2})\.(\d{1,2})", label)
    if not match:
        return None
    month = int(match.group(1))
    day = int(match.group(2))
    return f"{year}-{month:02d}-{day:02d}"


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def parse():
    wb = openpyxl.load_workbook(INPUT, data_only=True)
    months = {}
    latest = None
    member_index = {}

    for order, ws in enumerate(wb.worksheets, start=1):
        year, month = sheet_year_month(ws.title, order)
        if month < 1 or month > 12:
            continue

        missions = []
        for col in range(1, ws.max_column + 1):
            key = mission_key(year, ws.cell(1, col).value)
            if not key:
                continue
            missions.append(
                {
                    "date": key,
                    "label": clean(ws.cell(1, col).value),
                    "hashtag": clean(ws.cell(2, col).value),
                    "completedText": clean(ws.cell(3, col).value),
                    "col": col,
                }
            )

        if not missions:
            continue

        members = []
        for row in range(5, ws.max_row + 1):
            member_no = ws.cell(row, 2).value
            nickname = clean(ws.cell(row, 3).value)
            instagram_id = clean(ws.cell(row, 4).value)
            if not nickname and not instagram_id:
                continue
            if not isinstance(member_no, (int, float)):
                continue

            checks = {}
            for mission in missions:
                status = clean(ws.cell(row, mission["col"]).value)
                if status == "완료":
                    checks[mission["date"]] = True

            member_id = instagram_id or f"member-{int(member_no)}"
            member_index.setdefault(
                member_id,
                {
                    "id": member_id,
                    "emoji": "🙂",
                    "nickname": nickname or member_id,
                    "instagramId": instagram_id,
                },
            )
            if nickname:
                member_index[member_id]["nickname"] = nickname

            members.append(
                {
                    "id": member_id,
                    "memberNo": int(member_no),
                    "nickname": nickname or member_id,
                    "instagramId": instagram_id,
                    "checks": checks,
                }
            )

        key = f"{year}-{month}"
        months[key] = {
            "year": year,
            "month": month,
            "missions": [{k: v for k, v in mission.items() if k != "col"} for mission in missions],
            "members": members,
        }
        if latest is None:
            latest = {"year": year, "month": month}

    payload = {
        "source": INPUT.name,
        "latest": latest,
        "members": list(member_index.values()),
        "months": months,
    }
    OUTPUT.write_text(
        "window.DC_IMPORTED_DATA = "
        + json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
        + ";\n",
        encoding="utf-8",
    )
    print(json.dumps({"months": len(months), "members": len(member_index), "output": str(OUTPUT)}, ensure_ascii=False))


if __name__ == "__main__":
    parse()
